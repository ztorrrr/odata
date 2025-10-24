"""
OData v4 API Router
"""
import logging
from typing import Optional
from datetime import datetime
from io import StringIO, BytesIO
from pathlib import Path
from fastapi import APIRouter, Query, Request, Response, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import httpx
import tempfile
import os

from app.services.bigquery_service import get_bigquery_service
from app.services.odata_metadata import ODataMetadataGenerator
from app.services.odata_query_parser import ODataQueryParser
from app.services.excel_connection_modifier import ExcelConnectionModifier
from app.services.excel_template_generator import ExcelTemplateGenerator
from app.services.excel_with_connection import ExcelWithConnectionGenerator
from app.services.web_query_excel_generator import WebQueryExcelGenerator
from app.services.excel_with_pq_generator import ExcelWithPowerQueryGenerator
from app.utils.setting import get_config

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/odata", tags=["OData"])

# 서비스 인스턴스
config = get_config()
bq_service = get_bigquery_service()
metadata_generator = ODataMetadataGenerator()
query_parser = ODataQueryParser()


@router.get("/")
async def get_service_document():
    """
    OData Service Document
    Excel에서 OData 피드를 연결할 때 처음 호출되는 엔드포인트
    """
    logger.info("Service document requested")
    service_doc = metadata_generator.get_service_document()

    return JSONResponse(
        content=service_doc,
        headers={
            "OData-Version": "4.0",
            "Content-Type": "application/json;odata.metadata=minimal;charset=utf-8"
        }
    )


@router.get("/$metadata")
async def get_metadata():
    """
    OData Metadata Document
    데이터 구조 및 스키마 정보를 제공하는 XML 문서
    """
    logger.info("Metadata requested")
    metadata_xml = metadata_generator.generate_metadata()

    return Response(
        content=metadata_xml,
        media_type="application/xml",
        headers={
            "OData-Version": "4.0",
            "Content-Type": "application/xml; charset=utf-8"
        }
    )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}")
async def get_entity_set(
    request: Request,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
    top: Optional[int] = Query(None, alias="$top", description="Number of records to return", ge=1, le=10000),
    skip: Optional[int] = Query(None, alias="$skip", description="Number of records to skip", ge=0),
    count: Optional[bool] = Query(False, alias="$count", description="Include count of items"),
):
    """
    OData Entity Set Query
    테이블 데이터를 쿼리하는 메인 엔드포인트

    지원하는 쿼리 옵션:
    - $filter: 필터 조건 (예: name eq 'John')
    - $select: 선택할 필드 (예: name,age,email)
    - $orderby: 정렬 조건 (예: name desc)
    - $top: 최대 행 수 (예: 100)
    - $skip: 건너뛸 행 수 (예: 50)
    - $count: 전체 개수 포함 여부 (예: true)
    """
    try:
        logger.info(f"Entity set query: filter={filter}, select={select}, orderby={orderby}, top={top}, skip={skip}, count={count}")

        # 쿼리 파라미터 파싱
        parsed_params = query_parser.parse_all({
            "$filter": filter,
            "$select": select,
            "$orderby": orderby,
            "$top": str(top) if top else None,
            "$skip": str(skip) if skip else None,
            "$count": str(count)
        })

        # 기본값 설정
        if not parsed_params["top"]:
            parsed_params["top"] = config.ODATA_MAX_PAGE_SIZE

        # BigQuery 쿼리 실행
        rows = bq_service.query_table(
            select=parsed_params["select"],
            filter=parsed_params["filter"],
            orderby=parsed_params["orderby"],
            top=parsed_params["top"],
            skip=parsed_params["skip"]
        )

        # OData 응답 형식으로 변환
        base_url = str(request.base_url).rstrip('/')
        response_data = {
            "@odata.context": f"{base_url}/odata/$metadata#{config.BIGQUERY_TABLE_NAME}",
            "value": rows
        }

        # Count 포함
        if parsed_params["count"]:
            total_count = bq_service.get_row_count(filter=parsed_params["filter"])
            response_data["@odata.count"] = total_count

        # NextLink 추가 (페이징)
        if len(rows) == parsed_params["top"]:
            next_skip = (parsed_params["skip"] or 0) + parsed_params["top"]
            query_string = f"?$skip={next_skip}&$top={parsed_params['top']}"

            if filter:
                query_string += f"&$filter={filter}"
            if select:
                query_string += f"&$select={select}"
            if orderby:
                query_string += f"&$orderby={orderby}"

            response_data["@odata.nextLink"] = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}{query_string}"

        logger.info(f"Returning {len(rows)} rows")
        return JSONResponse(
            content=response_data,
            headers={
                "OData-Version": "4.0",
                "Content-Type": "application/json; odata.metadata=minimal"
            }
        )

    except Exception as e:
        logger.error(f"Error querying entity set: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/$count")
async def get_count(
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
):
    """
    OData Count Request
    필터된 결과의 개수만 반환
    """
    try:
        logger.info(f"Count requested with filter: {filter}")

        # 쿼리 파라미터 파싱
        parsed_filter = query_parser.parse_filter(filter)

        # 개수 조회
        count = bq_service.get_row_count(filter=parsed_filter)

        logger.info(f"Count result: {count}")

        return Response(
            content=str(count),
            media_type="text/plain",
            headers={
                "OData-Version": "4.0"
            }
        )

    except Exception as e:
        logger.error(f"Error getting count: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get("/health")
async def health_check():
    """
    헬스 체크 엔드포인트
    """
    try:
        # BigQuery 테이블 정보 조회로 연결 확인
        table_info = bq_service.get_table_info()

        if table_info:
            return {
                "status": "healthy",
                "bigquery": {
                    "connected": True,
                    "table": table_info["table_id"],
                    "rows": table_info["num_rows"]
                }
            }
        else:
            return JSONResponse(
                status_code=503,
                content={
                    "status": "unhealthy",
                    "error": "BigQuery table not found"
                }
            )

    except Exception as e:
        logger.error(f"Health check failed: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=503,
            content={
                "status": "unhealthy",
                "error": str(e)
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/export")
async def export_to_csv(
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
    top: Optional[int] = Query(100000, alias="$top", description="Maximum number of records to export", ge=1, le=1000000),
    skip: Optional[int] = Query(None, alias="$skip", description="Number of records to skip", ge=0),
):
    """
    CSV Export Endpoint
    테이블 데이터를 CSV 파일로 다운로드

    지원하는 쿼리 옵션:
    - $filter: 필터 조건 (예: name eq 'John')
    - $select: 선택할 필드 (예: name,age,email)
    - $orderby: 정렬 조건 (예: name desc)
    - $top: 최대 행 수 (기본: 100,000, 최대: 1,000,000)
    - $skip: 건너뛸 행 수 (예: 50)
    """
    try:
        logger.info(f"CSV export requested: filter={filter}, select={select}, orderby={orderby}, top={top}, skip={skip}")

        # 쿼리 파라미터 파싱
        parsed_params = query_parser.parse_all({
            "$filter": filter,
            "$select": select,
            "$orderby": orderby,
            "$top": str(top) if top else None,
            "$skip": str(skip) if skip else None,
            "$count": "false"
        })

        # BigQuery 쿼리 실행
        rows = bq_service.query_table(
            select=parsed_params["select"],
            filter=parsed_params["filter"],
            orderby=parsed_params["orderby"],
            top=parsed_params["top"],
            skip=parsed_params["skip"]
        )

        if not rows:
            return JSONResponse(
                status_code=404,
                content={
                    "error": {
                        "code": "NoData",
                        "message": "No data found for the given query"
                    }
                }
            )

        # pandas DataFrame으로 변환
        df = pd.DataFrame(rows)

        # CSV 생성 (UTF-8 BOM 추가로 Excel 호환성 확보)
        csv_buffer = StringIO()
        df.to_csv(csv_buffer, index=False, encoding='utf-8-sig')
        csv_content = csv_buffer.getvalue()

        # 파일명 생성 (테이블명_날짜.csv)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_{timestamp}.csv"

        logger.info(f"Exporting {len(rows)} rows to CSV file: {filename}")

        # StreamingResponse로 파일 다운로드 제공
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )

    except Exception as e:
        logger.error(f"Error exporting to CSV: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/connection")
async def get_odc_connection(
    request: Request,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
):
    """
    ODC (Office Data Connection) 파일 생성
    Excel에서 OData 연결을 바로 사용할 수 있는 연결 파일 제공

    - 데이터가 아닌 연결 정보만 포함 (파일 크기: ~1KB)
    - Excel에서 더블클릭으로 즉시 OData 연결 생성
    - 필요할 때마다 "새로고침"으로 최신 데이터 로드

    지원하는 쿼리 옵션:
    - $filter: 필터 조건 (연결에 포함)
    - $select: 선택할 필드 (연결에 포함)
    - $orderby: 정렬 조건 (연결에 포함)
    """
    try:
        logger.info(f"ODC connection requested: filter={filter}, select={select}, orderby={orderby}")

        # OData URL 구성
        base_url = str(request.base_url).rstrip('/')
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        # 쿼리 파라미터 추가
        query_params = []
        if filter:
            query_params.append(f"$filter={filter}")
        if select:
            query_params.append(f"$select={select}")
        if orderby:
            query_params.append(f"$orderby={orderby}")

        if query_params:
            odata_url += "?" + "&".join(query_params)

        # Power Query M 코드 생성
        m_code = f"""let
    원본 = OData.Feed("{odata_url}", null, [Implementation="2.0"])
in
    원본"""

        # ODC XML 생성 (Power Query 방식)
        # 핵심: 원본 구조 유지하되 HTML body를 최소화하여 자동 실행 방지
        # ODCDataSource 클래스와 JavaScript init()가 자동 실행의 주요 트리거
        odc_content = f"""<html xmlns:o="urn:schemas-microsoft-com:office:office"
xmlns="http://www.w3.org/TR/REC-html40">

<head>
<meta http-equiv=Content-Type content="text/x-ms-odc; charset=utf-8">
<meta name=ProgId content=ODC.Database>
<meta name=SourceType content=OLEDB>
<title>쿼리 - {config.BIGQUERY_TABLE_NAME}</title>
<xml id=docprops><o:DocumentProperties
  xmlns:o="urn:schemas-microsoft-com:office:office"
  xmlns="http://www.w3.org/TR/REC-html40">
  <o:Description>OData 연결: {config.BIGQUERY_TABLE_NAME}. 쿼리 및 연결에서 수동으로 로드하세요.</o:Description>
  <o:Name>쿼리 - {config.BIGQUERY_TABLE_NAME}</o:Name>
 </o:DocumentProperties>
</xml><xml id=msodc><odc:OfficeDataConnection
  xmlns:odc="urn:schemas-microsoft-com:office:odc"
  xmlns="http://www.w3.org/TR/REC-html40">
  <odc:PowerQueryConnection odc:Type="OLEDB">
   <odc:ConnectionString>Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;Location={config.BIGQUERY_TABLE_NAME};Extended Properties=</odc:ConnectionString>
   <odc:CommandType>TableCollection</odc:CommandType>
   <odc:CommandText>&quot;{config.BIGQUERY_TABLE_NAME}&quot;</odc:CommandText>
  </odc:PowerQueryConnection>
  <odc:PowerQueryMashupData>&lt;Mashup xmlns:xsd=&quot;http://www.w3.org/2001/XMLSchema&quot; xmlns:xsi=&quot;http://www.w3.org/2001/XMLSchema-instance&quot; xmlns=&quot;http://schemas.microsoft.com/DataMashup&quot;&gt;&lt;Client&gt;EXCEL&lt;/Client&gt;&lt;Version&gt;2.147.503.0&lt;/Version&gt;&lt;MinVersion&gt;2.21.0.0&lt;/MinVersion&gt;&lt;Culture&gt;ko-KR&lt;/Culture&gt;&lt;SafeCombine&gt;true&lt;/SafeCombine&gt;&lt;Items&gt;&lt;Query Name=&quot;{config.BIGQUERY_TABLE_NAME}&quot;&gt;&lt;Formula&gt;&lt;![CDATA[{m_code}]]&gt;&lt;/Formula&gt;&lt;IsParameterQuery xsi:nil=&quot;true&quot; /&gt;&lt;IsDirectQuery xsi:nil=&quot;true&quot; /&gt;&lt;/Query&gt;&lt;/Items&gt;&lt;/Mashup&gt;</odc:PowerQueryMashupData>
 </odc:OfficeDataConnection>
</xml>
</head>

<body>
<p style="font-family: Arial; font-size: 12px; padding: 20px;">
<strong>OData 연결 정보</strong><br/>
테이블: {config.BIGQUERY_TABLE_NAME}<br/>
URL: {odata_url}<br/><br/>
<strong>데이터 로드 방법:</strong><br/>
1. Excel의 '데이터' 탭을 클릭하세요<br/>
2. '쿼리 및 연결'을 클릭하세요<br/>
3. '{config.BIGQUERY_TABLE_NAME}' 쿼리를 마우스 오른쪽 클릭하세요<br/>
4. '로드 대상...'을 선택하세요<br/>
5. 원하는 위치를 선택하고 '로드'를 클릭하세요
</p>
</body>

</html>"""

        filename = f"{config.BIGQUERY_TABLE_NAME}_connection.odc"

        logger.info(f"Generating ODC file: {filename} for URL: {odata_url}")

        return Response(
            content=odc_content,
            media_type="application/x-ms-odc",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/x-ms-odc; charset=utf-8"
            }
        )

    except Exception as e:
        logger.error(f"Error generating ODC connection: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/template")
async def get_excel_template(
    request: Request,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    sample: Optional[int] = Query(10, alias="sample", description="Number of sample rows to include", ge=0, le=100),
):
    """
    Excel 템플릿 파일 생성 (OData 연결 설정 안내 포함)

    다운로드 후 바로 사용 가능한 Excel 파일 제공:
    - 첫 번째 시트: OData 연결 방법 안내
    - 두 번째 시트: Power Query M 코드 (복사/붙여넣기용)
    - 세 번째 시트: 샘플 데이터 (preview)

    지원하는 파라미터:
    - $filter: 필터 조건
    - $select: 선택할 필드
    - sample: 샘플 데이터 행 수 (기본: 10)
    """
    try:
        logger.info(f"Excel template requested: filter={filter}, select={select}, sample={sample}")

        # OData URL 구성
        base_url = str(request.base_url).rstrip('/')
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        query_params = []
        if filter:
            query_params.append(f"$filter={filter}")
        if select:
            query_params.append(f"$select={select}")

        if query_params:
            odata_url += "?" + "&".join(query_params)

        # Excel 워크북 생성
        wb = Workbook()

        # === 첫 번째 시트: 사용 방법 안내 ===
        ws_guide = wb.active
        ws_guide.title = "사용 방법"

        # 제목
        ws_guide['A1'] = f"{config.BIGQUERY_TABLE_NAME} OData 연결 템플릿"
        ws_guide['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws_guide['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws_guide['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws_guide.merge_cells('A1:D1')
        ws_guide.row_dimensions[1].height = 30

        # 안내 내용
        guide_content = [
            ["", "", "", ""],
            ["OData URL:", odata_url, "", ""],
            ["", "", "", ""],
            ["Excel에서 OData 데이터 연결하는 방법:", "", "", ""],
            ["", "", "", ""],
            ["방법 1) Power Query 사용 (추천)", "", "", ""],
            ["  1. 데이터 탭 → 데이터 가져오기 → OData 피드에서", "", "", ""],
            ["  2. 위의 'OData URL' 복사하여 붙여넣기", "", "", ""],
            ["  3. 익명 인증 선택 → 확인", "", "", ""],
            ["  4. 데이터 로드", "", "", ""],
            ["", "", "", ""],
            ["방법 2) M 코드 사용 (고급)", "", "", ""],
            ["  1. 데이터 탭 → 데이터 가져오기 → 빈 쿼리", "", "", ""],
            ["  2. 홈 탭 → 고급 편집기", "", "", ""],
            ["  3. 'M 코드' 시트의 코드 복사/붙여넣기", "", "", ""],
            ["  4. 완료 → 로드", "", "", ""],
            ["", "", "", ""],
            ["방법 3) ODC 파일 사용", "", "", ""],
            ["  1. 다음 URL에서 ODC 파일 다운로드:", "", "", ""],
            [f"     {base_url}/odata/{config.BIGQUERY_TABLE_NAME}/connection", "", "", ""],
            ["  2. 파일 더블클릭 → 가져오기", "", "", ""],
            ["", "", "", ""],
            ["샘플 데이터는 'Sample Data' 시트를 참고하세요.", "", "", ""],
        ]

        for row_idx, row_data in enumerate(guide_content, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_guide.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1 and value and "방법" in value:
                    cell.font = Font(bold=True, size=11)

        # URL 셀 스타일
        ws_guide['B2'].font = Font(color="0563C1", underline="single")
        ws_guide['A20'].font = Font(size=9, color="0563C1")

        # 열 너비 조정
        ws_guide.column_dimensions['A'].width = 50
        ws_guide.column_dimensions['B'].width = 60

        # === 두 번째 시트: Power Query M 코드 ===
        ws_code = wb.create_sheet(title="M 코드")

        m_code = f'''let
    Source = OData.Feed(
        "{odata_url}",
        null,
        [Implementation="2.0"]
    )
in
    Source'''

        ws_code['A1'] = "Power Query M 코드"
        ws_code['A1'].font = Font(size=14, bold=True)
        ws_code['A3'] = "아래 코드를 복사하여 Power Query 고급 편집기에 붙여넣으세요:"
        ws_code['A3'].font = Font(italic=True)

        ws_code['A5'] = m_code
        ws_code['A5'].font = Font(name="Consolas", size=10)
        ws_code['A5'].alignment = Alignment(wrap_text=True, vertical="top")

        ws_code.column_dimensions['A'].width = 80
        ws_code.row_dimensions[5].height = 100

        # === 세 번째 시트: 샘플 데이터 ===
        if sample > 0:
            ws_sample = wb.create_sheet(title="Sample Data")

            # 쿼리 파라미터 파싱
            parsed_params = query_parser.parse_all({
                "$filter": filter,
                "$select": select,
                "$orderby": None,
                "$top": str(sample),
                "$skip": None,
                "$count": "false"
            })

            # BigQuery에서 샘플 데이터 조회
            rows = bq_service.query_table(
                select=parsed_params["select"],
                filter=parsed_params["filter"],
                orderby=None,
                top=sample,
                skip=None
            )

            if rows:
                df = pd.DataFrame(rows)

                # 헤더 추가
                ws_sample['A1'] = f"샘플 데이터 ({len(rows)}행)"
                ws_sample['A1'].font = Font(size=12, bold=True)
                last_col = get_column_letter(len(df.columns))
                ws_sample.merge_cells(f'A1:{last_col}1')

                # 데이터 추가
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
                    for c_idx, value in enumerate(row, start=1):
                        cell = ws_sample.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 3:  # 헤더 행
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

                # 자동 열 너비 조정
                for col_idx in range(1, len(df.columns) + 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)
                    for row in ws_sample.iter_rows(min_col=col_idx, max_col=col_idx, min_row=3):
                        for cell in row:
                            try:
                                if cell.value and not isinstance(cell, type(ws_sample['A1'])):
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_sample.column_dimensions[column_letter].width = adjusted_width

        # Excel 파일을 BytesIO로 저장
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_template_{timestamp}.xlsx"

        logger.info(f"Generated Excel template: {filename} with {sample} sample rows")

        return StreamingResponse(
            excel_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
            }
        )

    except Exception as e:
        logger.error(f"Error generating Excel template: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/excel-live")
async def get_excel_with_live_connection(
    request: Request,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
    sheet_name: Optional[str] = Query("Data", description="Sheet name in Excel"),
    query_name: Optional[str] = Query("ODataQuery", description="Query name in Excel Power Query"),
):
    """
    Excel 파일 생성 with Live OData Connection (Windows COM 방식)

    Windows Excel Service에 요청을 위임하여 실제 Power Query 연결이 포함된 Excel 파일 생성:
    - 파일을 열면 OData 데이터가 자동으로 로드됨
    - Excel에서 '새로고침' 버튼으로 최신 데이터 갱신 가능
    - Power Query M 코드가 내장되어 있어 연결 정보 보존

    ⚠️ 이 엔드포인트는 별도의 Windows 서버가 필요합니다.
    환경 변수 WINDOWS_EXCEL_SERVICE_URL 설정 필요.

    지원하는 파라미터:
    - $filter: 필터 조건
    - $select: 선택할 필드
    - $orderby: 정렬 조건
    - sheet_name: 시트 이름 (기본: Data)
    - query_name: 쿼리 이름 (기본: ODataQuery)
    """
    try:
        # Windows Excel Service URL 확인
        if not config.WINDOWS_EXCEL_SERVICE_URL:
            raise HTTPException(
                status_code=501,
                detail={
                    "error": "Windows Excel Service not configured",
                    "message": "WINDOWS_EXCEL_SERVICE_URL environment variable is not set",
                    "alternative": f"Use /{config.BIGQUERY_TABLE_NAME}/template or /{config.BIGQUERY_TABLE_NAME}/connection instead"
                }
            )

        logger.info(f"Excel live connection requested: filter={filter}, select={select}, orderby={orderby}")

        # OData URL 구성
        base_url = str(request.base_url).rstrip('/')
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        query_params = []
        if filter:
            query_params.append(f"$filter={filter}")
        if select:
            query_params.append(f"$select={select}")
        if orderby:
            query_params.append(f"$orderby={orderby}")

        if query_params:
            odata_url += "?" + "&".join(query_params)

        # Windows Excel Service에 요청 전달
        windows_service_url = f"{config.WINDOWS_EXCEL_SERVICE_URL.rstrip('/')}/excel/odata"

        logger.info(f"Delegating to Windows Excel Service: {windows_service_url}")
        logger.info(f"OData URL: {odata_url}")

        async with httpx.AsyncClient(timeout=config.WINDOWS_EXCEL_SERVICE_TIMEOUT) as client:
            response = await client.get(
                windows_service_url,
                params={
                    "odata_url": odata_url,
                    "sheet_name": sheet_name,
                    "query_name": query_name
                }
            )

            if response.status_code != 200:
                logger.error(f"Windows Excel Service error: {response.status_code} - {response.text}")
                raise HTTPException(
                    status_code=502,
                    detail={
                        "error": "Windows Excel Service error",
                        "status_code": response.status_code,
                        "message": response.text,
                        "service_url": config.WINDOWS_EXCEL_SERVICE_URL
                    }
                )

            # 임시 파일로 저장
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_file.write(response.content)
            temp_file.close()

            # 파일명 생성
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{config.BIGQUERY_TABLE_NAME}_live_{timestamp}.xlsx"

            logger.info(f"Excel file generated via Windows service: {filename}")

            # FileResponse로 반환 (background에서 임시 파일 삭제)
            return FileResponse(
                path=temp_file.name,
                filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                background=lambda: os.unlink(temp_file.name)
            )

    except httpx.TimeoutException:
        logger.error(f"Windows Excel Service timeout after {config.WINDOWS_EXCEL_SERVICE_TIMEOUT}s")
        return JSONResponse(
            status_code=504,
            content={
                "error": {
                    "code": "ServiceTimeout",
                    "message": f"Windows Excel Service did not respond within {config.WINDOWS_EXCEL_SERVICE_TIMEOUT} seconds",
                    "service_url": config.WINDOWS_EXCEL_SERVICE_URL
                }
            }
        )

    except httpx.ConnectError:
        logger.error(f"Cannot connect to Windows Excel Service: {config.WINDOWS_EXCEL_SERVICE_URL}")
        return JSONResponse(
            status_code=503,
            content={
                "error": {
                    "code": "ServiceUnavailable",
                    "message": "Cannot connect to Windows Excel Service",
                    "service_url": config.WINDOWS_EXCEL_SERVICE_URL,
                    "alternative": f"Use /{config.BIGQUERY_TABLE_NAME}/template or /{config.BIGQUERY_TABLE_NAME}/connection instead"
                }
            }
        )

    except HTTPException:
        raise

    except Exception as e:
        logger.error(f"Error requesting Excel with live connection: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/excel-template")
async def get_excel_from_template(
    request: Request,
    background_tasks: BackgroundTasks,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
):
    """
    템플릿 Excel 파일의 OData 연결 수정하여 제공

    템플릿 파일(app/template/odata_template.xlsx)의 Power Query 연결 정보를
    요청된 엔드포인트로 변경하여 제공합니다.

    특징:
    - 데이터를 시트에 포함하지 않음 (연결 정보만 저장)
    - 대용량 데이터도 처리 가능
    - 사용자는 파일 다운로드 후 Excel에서 '데이터 새로고침'만 하면 됨
    - 필터/선택/정렬 조건이 연결에 포함됨

    지원하는 파라미터:
    - $filter: 필터 조건 (예: Media eq 'Naver')
    - $select: 선택할 필드 (예: Date,Campaign,Clicks)
    - $orderby: 정렬 조건 (예: Date desc)
    """
    try:
        logger.info(f"Excel template requested: filter={filter}, select={select}, orderby={orderby}")

        # 템플릿 파일 경로
        template_path = Path(__file__).parent.parent / "template" / "odata_template.xlsx"

        if not template_path.exists():
            raise HTTPException(
                status_code=500,
                detail={
                    "error": "Template file not found",
                    "message": f"Template file does not exist: {template_path}",
                    "path": str(template_path)
                }
            )

        # OData URL 구성
        base_url = str(request.base_url).rstrip('/')
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        query_params = []
        if filter:
            query_params.append(f"$filter={filter}")
        if select:
            query_params.append(f"$select={select}")
        if orderby:
            query_params.append(f"$orderby={orderby}")

        if query_params:
            odata_url += "?" + "&".join(query_params)

        logger.info(f"Target OData URL: {odata_url}")

        # Excel 연결 수정기 초기화
        modifier = ExcelConnectionModifier(str(template_path))

        # 연결 정보 수정하여 새 파일 생성
        modified_file_path = modifier.modify_odata_connection(odata_url)

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_connection_{timestamp}.xlsx"

        logger.info(f"Generated Excel file from template: {filename}")

        # 임시 파일 삭제 작업을 background task로 추가
        background_tasks.add_task(os.unlink, modified_file_path)

        # FileResponse로 반환
        return FileResponse(
            path=modified_file_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise

    except Exception as e:
        logger.error(f"Error generating Excel from template: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/simple-excel")
async def get_simple_excel_template(
    request: Request,
    background_tasks: BackgroundTasks,
):
    """
    단순한 Excel 템플릿 제공 (Power Query 없이)

    OData 연결 방법을 안내하는 단순한 Excel 파일을 제공합니다.
    사용자가 수동으로 연결을 설정할 수 있도록 안내합니다.
    """
    try:
        # OData URL 구성
        base_url = str(request.base_url).rstrip("/")
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        # Excel 생성기 초기화
        generator = ExcelTemplateGenerator()

        # 단순 템플릿 생성
        output_path = generator.generate_simple_template(
            odata_url=odata_url,
            table_name=config.BIGQUERY_TABLE_NAME
        )

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_simple_{timestamp}.xlsx"

        # 임시 파일 삭제 작업 추가
        background_tasks.add_task(os.unlink, output_path)

        return FileResponse(
            path=output_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating simple Excel template: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/odc")
async def get_odc_connection_file(
    request: Request,
    background_tasks: BackgroundTasks,
):
    """
    ODC (Office Data Connection) 파일 제공

    Excel에서 직접 열 수 있는 연결 파일입니다.
    다운로드 후 더블클릭하면 Excel이 자동으로 OData 연결을 생성합니다.
    """
    try:
        # OData URL 구성
        base_url = str(request.base_url).rstrip("/")
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        # Excel 생성기 초기화
        generator = ExcelTemplateGenerator()

        # ODC 파일 생성
        output_path = generator.generate_odc_file(
            odata_url=odata_url,
            table_name=config.BIGQUERY_TABLE_NAME
        )

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_{timestamp}.odc"

        # 임시 파일 삭제 작업 추가
        background_tasks.add_task(os.unlink, output_path)

        return FileResponse(
            path=output_path,
            filename=filename,
            media_type="text/x-ms-odc"
        )

    except Exception as e:
        logger.error(f"Error generating ODC file: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/excel-connection")
async def get_excel_with_connection_only(
    request: Request,
    background_tasks: BackgroundTasks,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
):
    """
    Power Query 연결이 포함된 Excel 파일 생성 (데이터 미포함)

    "쿼리 및 연결" 패널에 연결 정보만 포함된 Excel 파일을 생성합니다.
    - 시트에는 데이터가 없고 사용 안내만 포함
    - Power Query 연결이 설정되어 있음
    - 사용자가 수동으로 로드하거나 새로고침 가능

    지원하는 파라미터:
    - $filter: 필터 조건
    - $select: 선택할 필드
    - $orderby: 정렬 조건
    """
    try:
        logger.info(f"Excel with connection requested: filter={filter}, select={select}, orderby={orderby}")

        # OData URL 구성
        base_url = str(request.base_url).rstrip('/')
        odata_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}"

        # 쿼리 파라미터
        query_params = {
            'filter': filter,
            'select': select,
            'orderby': orderby
        }

        # Excel 생성기 초기화
        generator = ExcelWithConnectionGenerator()

        # 연결이 포함된 Excel 파일 생성
        output_path = generator.generate_excel_with_connection(
            odata_url=odata_url,
            table_name=config.BIGQUERY_TABLE_NAME,
            query_params=query_params
        )

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_query_{timestamp}.xlsx"

        # 임시 파일 삭제 작업 추가
        background_tasks.add_task(os.unlink, output_path)

        return FileResponse(
            path=output_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating Excel with connection: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/json")
async def get_data_as_json(
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
    top: Optional[int] = Query(10000, alias="$top", description="Maximum number of rows", le=100000),
    skip: Optional[int] = Query(None, alias="$skip", description="Number of rows to skip"),
):
    """
    BigQuery 데이터를 JSON으로 반환 (Excel Power Query 연결용)

    OData v4 형식의 JSON 응답을 반환합니다.
    Excel Power Query에서 Web.Contents()로 직접 호출 가능합니다.

    지원하는 파라미터:
    - $filter: 필터 조건
    - $select: 선택할 필드
    - $orderby: 정렬 조건
    - $top: 최대 행 수 (기본: 10000, 최대: 100000)
    - $skip: 건너뛸 행 수
    """
    try:
        bq_service = get_bigquery_service()
        query_parser = ODataQueryParser()

        logger.info(f"JSON API request: filter={filter}, select={select}, top={top}")

        # OData 파라미터 파싱
        parsed_params = query_parser.parse_all({
            "$filter": filter,
            "$select": select,
            "$orderby": orderby,
            "$top": str(top) if top else None,
            "$skip": str(skip) if skip else None,
            "$count": "false"
        })

        # BigQuery 쿼리 실행
        rows = bq_service.query_table(
            select=parsed_params["select"],
            filter=parsed_params["filter"],
            orderby=parsed_params["orderby"],
            top=top,
            skip=skip
        )

        # OData v4 형식 응답
        response_data = {
            "@odata.context": f"$metadata#{config.BIGQUERY_TABLE_NAME}",
            "value": rows
        }

        # 카운트가 top과 일치하면 nextLink 추가
        if len(rows) == top:
            next_skip = (skip or 0) + top
            response_data["@odata.nextLink"] = f"/{config.BIGQUERY_TABLE_NAME}/json?$skip={next_skip}&$top={top}"
            if filter:
                response_data["@odata.nextLink"] += f"&$filter={filter}"
            if select:
                response_data["@odata.nextLink"] += f"&$select={select}"
            if orderby:
                response_data["@odata.nextLink"] += f"&$orderby={orderby}"

        logger.info(f"Returning {len(rows)} rows as JSON")

        return JSONResponse(content=response_data)

    except Exception as e:
        logger.error(f"Error querying data as JSON: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/web-query-excel")
async def get_web_query_excel(
    request: Request,
    background_tasks: BackgroundTasks,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
    top: Optional[int] = Query(10000, alias="$top", description="Maximum number of rows", le=100000),
):
    """
    웹 JSON API 연결이 포함된 Excel 파일 생성 (데이터 미포함)

    ✅ 안정적인 방법: Power Query M 코드로 JSON API 연결
    - 시트에 데이터 없음 (연결 정보만)
    - 사용자가 M 코드를 복사/붙여넣기하여 연결 설정
    - 새로고침으로 최신 데이터 로드

    장점:
    - 자동 실행 없음 (사용자가 제어)
    - 안정적이고 오류 없음
    - 빅데이터 처리 가능
    - OData Power Query 손상 문제 우회

    지원하는 파라미터:
    - $filter: 필터 조건
    - $select: 선택할 필드
    - $orderby: 정렬 조건
    - $top: 최대 행 수 (기본: 10000)
    """
    try:
        logger.info(f"Web Query Excel requested: filter={filter}, select={select}, top={top}")

        # JSON API URL 구성
        base_url = str(request.base_url).rstrip('/')
        json_api_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}/json"

        # 쿼리 파라미터 추가
        query_params = []
        if filter:
            query_params.append(f"$filter={filter}")
        if select:
            query_params.append(f"$select={select}")
        if orderby:
            query_params.append(f"$orderby={orderby}")
        if top:
            query_params.append(f"$top={top}")

        if query_params:
            json_api_url += "?" + "&".join(query_params)

        # 설명 생성
        description = f"BigQuery 테이블 '{config.BIGQUERY_TABLE_NAME}'의 JSON API 연결"
        if filter:
            description += f" (필터: {filter})"

        # Excel 생성기 초기화
        generator = WebQueryExcelGenerator()

        # JSON 연결 Excel 파일 생성
        output_path = generator.generate_json_connection_excel(
            json_api_url=json_api_url,
            table_name=config.BIGQUERY_TABLE_NAME,
            description=description
        )

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_webquery_{timestamp}.xlsx"

        logger.info(f"Generated web query Excel: {filename}")

        # 임시 파일 삭제 작업 추가
        background_tasks.add_task(os.unlink, output_path)

        return FileResponse(
            path=output_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating web query Excel: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )


@router.get(f"/{config.BIGQUERY_TABLE_NAME}/excel-ready")
async def get_excel_with_ready_connection(
    request: Request,
    background_tasks: BackgroundTasks,
    filter: Optional[str] = Query(None, alias="$filter", description="OData filter expression"),
    select: Optional[str] = Query(None, alias="$select", description="Comma-separated list of properties"),
    orderby: Optional[str] = Query(None, alias="$orderby", description="Order by expression"),
    top: Optional[int] = Query(10000, alias="$top", description="Maximum number of rows", le=100000),
):
    """
    Power Query 연결이 이미 설정된 Excel 파일 생성 ⭐ 권장

    ✅ 이 엔드포인트는 연결이 완전히 설정된 Excel 파일을 제공합니다:
    - Power Query 연결이 "쿼리 및 연결"에 표시됨
    - 시트에 데이터 없음 (연결 정보만)
    - 사용자가 "새로고침" 또는 "로드 대상"으로 데이터 로드
    - 간단한 DataMashup 사용 (OData 손상 문제 없음)
    - JSON API 사용 (안정적)

    사용자 워크플로우:
    1. 파일 다운로드
    2. Excel에서 파일 열기
    3. 데이터 탭 → "쿼리 및 연결"에서 쿼리 확인
    4. 쿼리 → "로드 대상..." 또는 "모두 새로고침" 클릭
    5. 데이터 로드 완료

    지원하는 파라미터:
    - $filter: 필터 조건
    - $select: 선택할 필드
    - $orderby: 정렬 조건
    - $top: 최대 행 수 (기본: 10000, 최대: 100000)
    """
    try:
        logger.info(f"Excel with ready connection requested: filter={filter}, select={select}, top={top}")

        # JSON API URL 구성
        base_url = str(request.base_url).rstrip('/')
        json_api_url = f"{base_url}/odata/{config.BIGQUERY_TABLE_NAME}/json"

        # 쿼리 파라미터 추가
        query_params = []
        if filter:
            query_params.append(f"$filter={filter}")
        if select:
            query_params.append(f"$select={select}")
        if orderby:
            query_params.append(f"$orderby={orderby}")
        if top:
            query_params.append(f"$top={top}")

        if query_params:
            json_api_url += "?" + "&".join(query_params)

        # Excel 생성기 초기화
        generator = ExcelWithPowerQueryGenerator()

        # Power Query 연결이 설정된 Excel 파일 생성
        output_path = generator.generate_excel_with_power_query(
            json_api_url=json_api_url,
            table_name=config.BIGQUERY_TABLE_NAME,
            query_name=config.BIGQUERY_TABLE_NAME
        )

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{config.BIGQUERY_TABLE_NAME}_ready_{timestamp}.xlsx"

        logger.info(f"Generated Excel with ready connection: {filename}")

        # 임시 파일 삭제 작업 추가
        background_tasks.add_task(os.unlink, output_path)

        return FileResponse(
            path=output_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        logger.error(f"Error generating Excel with ready connection: {str(e)}", exc_info=True)
        return JSONResponse(
            status_code=500,
            content={
                "error": {
                    "code": "InternalServerError",
                    "message": str(e)
                }
            }
        )
