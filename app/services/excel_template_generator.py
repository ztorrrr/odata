"""
Excel Template Generator
OData 연결이 포함된 Excel 템플릿을 생성
"""
import logging
import tempfile
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl import Workbook

logger = logging.getLogger(__name__)


class ExcelTemplateGenerator:
    """
    OData 연결 정보를 포함한 Excel 템플릿을 처음부터 생성
    """

    def generate_simple_template(
        self,
        odata_url: str,
        table_name: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        단순한 Excel 템플릿 생성 (Power Query 없이 OData 연결 정보만 포함)

        Args:
            odata_url: OData 서비스 URL
            table_name: 테이블 이름
            output_path: 출력 경로

        Returns:
            생성된 파일 경로
        """
        try:
            # 출력 파일 경로 결정
            if output_path is None:
                output_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".xlsx"
                )
                output_path = output_file.name
                output_file.close()

            # 새 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "OData Connection Info"

            # 헤더 추가
            ws['A1'] = "OData Connection Information"
            ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

            ws['A3'] = "Service URL:"
            ws['B3'] = odata_url

            ws['A4'] = "Table Name:"
            ws['B4'] = table_name

            ws['A6'] = "Instructions:"
            ws['A7'] = "1. Go to Data tab → Get Data → From Other Sources → From OData Feed"
            ws['A8'] = f"2. Enter URL: {odata_url}"
            ws['A9'] = "3. Select authentication method (usually Anonymous for local testing)"
            ws['A10'] = "4. Click Connect"
            ws['A11'] = "5. Select the table and click Load"

            # Power Query M 코드 제공
            ws['A13'] = "Alternative: Power Query M Code"
            ws['A13'].font = openpyxl.styles.Font(bold=True)

            ws['A14'] = "You can also use this M code in Power Query Advanced Editor:"

            m_code = f'''let
    Source = OData.Feed("{odata_url}", null, [Implementation="2.0"])
in
    Source'''

            ws['A16'] = m_code
            ws['A16'].alignment = openpyxl.styles.Alignment(wrap_text=True)

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 80
            ws.column_dimensions['B'].width = 60

            # 파일 저장
            wb.save(output_path)

            logger.info(f"Generated simple Excel template: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel template: {str(e)}", exc_info=True)
            raise

    def generate_odc_file(
        self,
        odata_url: str,
        table_name: str,
        output_path: Optional[str] = None
    ) -> str:
        """
        ODC (Office Data Connection) 파일 생성
        Excel에서 직접 열 수 있는 연결 파일

        Args:
            odata_url: OData 서비스 URL
            table_name: 테이블 이름
            output_path: 출력 경로

        Returns:
            생성된 ODC 파일 경로
        """
        try:
            if output_path is None:
                output_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".odc"
                )
                output_path = output_file.name
                output_file.close()

            # ODC 파일 내용 생성
            # 핵심: HTML, JavaScript 제거, 순수 연결 정보만 포함
            # PowerQueryConnection에서 연결 속성만 정의하고 자동 실행 방지
            odc_content = f'''<html xmlns:o="urn:schemas-microsoft-com:office:office">
<head>
<meta http-equiv=Content-Type content="text/x-ms-odc; charset=utf-8">
<meta name=ProgId content=ODC.Database>
<meta name=SourceType content=OLEDB>
<title>{table_name} - OData Connection (Manual Load Only)</title>
<xml id=docprops><o:DocumentProperties
  xmlns:o="urn:schemas-microsoft-com:office:office">
  <o:Description>OData connection to {table_name}. Click 'Load To' in Queries panel to load data.</o:Description>
  <o:Name>{table_name}</o:Name>
 </o:DocumentProperties>
</xml>
<xml id=msodc><odc:OfficeDataConnection
  xmlns:odc="urn:schemas-microsoft-com:office:odc">
  <odc:ConnectionString></odc:ConnectionString>
  <odc:CommandType>Default</odc:CommandType>
  <odc:CommandText></odc:CommandText>
  <odc:PowerQueryConnection odc:Type="Mashup">
   <odc:RefreshInfo>
    <odc:RefreshOnFileOpen>false</odc:RefreshOnFileOpen>
    <odc:Background>false</odc:Background>
    <odc:EnableRefresh>true</odc:EnableRefresh>
    <odc:SaveData>false</odc:SaveData>
   </odc:RefreshInfo>
  </odc:PowerQueryConnection>
  <odc:PowerQueryMashupData>&lt;Mashup xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns="http://schemas.microsoft.com/DataMashup"&gt;&lt;Client&gt;EXCEL&lt;/Client&gt;&lt;Version&gt;2.116.622.0&lt;/Version&gt;&lt;MinVersion&gt;2.21.0.0&lt;/MinVersion&gt;&lt;Culture&gt;en-US&lt;/Culture&gt;&lt;SafeCombine&gt;false&lt;/SafeCombine&gt;&lt;Items&gt;&lt;Query Name="{table_name}"&gt;&lt;Formula&gt;&lt;![CDATA[let
    Source = OData.Feed("{odata_url}", null, [Implementation="2.0"])
in
    Source]]&gt;&lt;/Formula&gt;&lt;IsParameterQuery xsi:nil="true" /&gt;&lt;IsDirectQuery xsi:nil="true" /&gt;&lt;/Query&gt;&lt;/Items&gt;&lt;/Mashup&gt;</odc:PowerQueryMashupData>
 </odc:OfficeDataConnection>
</xml>
</head>
<body>
</body>
</html>'''

            # ODC 파일 저장
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(odc_content)

            logger.info(f"Generated ODC file: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating ODC file: {str(e)}", exc_info=True)
            raise