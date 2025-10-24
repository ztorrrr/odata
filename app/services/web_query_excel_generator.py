"""
Web Query Excel Generator
웹 데이터(JSON API)에 연결된 Excel 파일 생성
"""
import logging
import tempfile
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


class WebQueryExcelGenerator:
    """
    웹 데이터 소스(JSON API)에 연결된 Excel 파일 생성
    Power Query M 코드를 포함하여 사용자가 쉽게 연결 설정 가능
    """

    def generate_json_connection_excel(
        self,
        json_api_url: str,
        table_name: str,
        description: Optional[str] = None,
        output_path: Optional[str] = None
    ) -> str:
        """
        JSON API 연결이 포함된 Excel 파일 생성

        Args:
            json_api_url: JSON API 엔드포인트 URL
            table_name: 테이블/쿼리 이름
            description: 연결 설명
            output_path: 출력 경로

        Returns:
            생성된 Excel 파일 경로
        """
        try:
            if output_path is None:
                output_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".xlsx"
                )
                output_path = output_file.name
                output_file.close()

            # Power Query M 코드 생성
            m_code = f'''let
    // JSON API 호출
    Source = Json.Document(Web.Contents("{json_api_url}")),

    // OData 형식의 응답 처리
    value = if Record.HasFields(Source, "value") then Source[value] else Source,

    // List를 Table로 변환
    ConvertedToTable = if value is list then Table.FromList(value, Splitter.SplitByNothing(), null, null, ExtraValues.Error) else Table.FromRecords(value),

    // Record 컬럼 확장 (필요시)
    ExpandedTable = if Table.ColumnNames(ConvertedToTable) = {{"Column1"}} then
        Table.ExpandRecordColumn(
            ConvertedToTable,
            "Column1",
            Record.FieldNames(ConvertedToTable{{0}}[Column1]),
            Record.FieldNames(ConvertedToTable{{0}}[Column1])
        )
    else
        ConvertedToTable
in
    ExpandedTable'''

            # Excel 워크북 생성
            wb = Workbook()

            # === 첫 번째 시트: 사용 방법 안내 ===
            ws_guide = wb.active
            ws_guide.title = "연결 설정 방법"

            # 제목
            ws_guide['A1'] = f"{table_name} - 웹 데이터 연결"
            ws_guide['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws_guide['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws_guide.merge_cells('A1:B1')
            ws_guide.row_dimensions[1].height = 30

            # 연결 정보
            ws_guide['A3'] = "데이터 소스:"
            ws_guide['A3'].font = Font(bold=True)
            ws_guide['B3'] = json_api_url

            if description:
                ws_guide['A4'] = "설명:"
                ws_guide['A4'].font = Font(bold=True)
                ws_guide['B4'] = description

            # 사용 방법
            ws_guide['A6'] = "📋 연결 설정 방법 (2분 소요)"
            ws_guide['A6'].font = Font(size=12, bold=True)

            instructions = [
                ("1단계", "Excel 데이터 탭 클릭"),
                ("2단계", "데이터 가져오기 → 기타 원본에서 → 빈 쿼리"),
                ("3단계", "Power Query 창에서 '고급 편집기' 클릭"),
                ("4단계", "기존 코드를 모두 삭제"),
                ("5단계", "'M 코드' 시트의 코드를 복사하여 붙여넣기"),
                ("6단계", "완료 → 닫기 및 로드"),
                ("", ""),
                ("✅ 완료!", "이후 '새로고침' 버튼으로 최신 데이터 로드 가능"),
            ]

            row = 8
            for step, desc in instructions:
                ws_guide[f'A{row}'] = step
                ws_guide[f'B{row}'] = desc
                if step:
                    ws_guide[f'A{row}'].font = Font(bold=True)
                row += 1

            # 주의사항
            ws_guide[f'A{row + 1}'] = "⚠️ 중요사항"
            ws_guide[f'A{row + 1}'].font = Font(size=11, bold=True, color="FF0000")

            ws_guide[f'A{row + 2}'] = "• 이 파일에는 데이터가 포함되어 있지 않습니다"
            ws_guide[f'A{row + 3}'] = "• 연결 설정 후 '새로고침' 버튼으로 데이터를 로드하세요"
            ws_guide[f'A{row + 4}'] = "• 대용량 데이터의 경우 필터 조건을 URL에 추가하세요"

            # 컬럼 너비
            ws_guide.column_dimensions['A'].width = 15
            ws_guide.column_dimensions['B'].width = 70

            # === 두 번째 시트: M 코드 ===
            ws_code = wb.create_sheet(title="M 코드")

            ws_code['A1'] = "Power Query M 코드 (전체 선택하여 복사)"
            ws_code['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws_code['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            ws_code.merge_cells('A1:A2')

            ws_code['A4'] = m_code
            ws_code['A4'].font = Font(name="Consolas", size=10)
            ws_code['A4'].alignment = Alignment(wrap_text=True, vertical="top")
            ws_code['A4'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            ws_code.column_dimensions['A'].width = 100
            ws_code.row_dimensions[4].height = 200

            # === 세 번째 시트: API 정보 ===
            ws_api = wb.create_sheet(title="API 정보")

            ws_api['A1'] = "API 엔드포인트 정보"
            ws_api['A1'].font = Font(size=14, bold=True)

            ws_api['A3'] = "URL:"
            ws_api['A3'].font = Font(bold=True)
            ws_api['B3'] = json_api_url

            ws_api['A5'] = "응답 형식:"
            ws_api['A5'].font = Font(bold=True)
            ws_api['B5'] = "JSON (OData v4 형식)"

            ws_api['A7'] = "예제 응답 구조:"
            ws_api['A7'].font = Font(bold=True)
            ws_api['A8'] = '''{
  "@odata.context": "...",
  "value": [
    {
      "field1": "value1",
      "field2": "value2",
      ...
    }
  ]
}'''
            ws_api['A8'].font = Font(name="Consolas", size=9)
            ws_api['A8'].alignment = Alignment(wrap_text=True, vertical="top")

            ws_api['A15'] = "필터 사용 예:"
            ws_api['A15'].font = Font(bold=True)
            ws_api['A16'] = f"{json_api_url}?$filter=Media eq 'Naver'"
            ws_api['A17'] = f"{json_api_url}?$select=Date,Campaign,Clicks"
            ws_api['A18'] = f"{json_api_url}?$top=1000"

            ws_api.column_dimensions['A'].width = 80
            ws_api.column_dimensions['B'].width = 50

            # 파일 저장
            wb.save(output_path)

            logger.info(f"Generated JSON connection Excel file: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating JSON connection Excel: {str(e)}", exc_info=True)
            raise


    def generate_csv_connection_excel(
        self,
        csv_url: str,
        table_name: str,
        has_headers: bool = True,
        delimiter: str = ",",
        output_path: Optional[str] = None
    ) -> str:
        """
        CSV 파일 연결이 포함된 Excel 파일 생성

        Args:
            csv_url: CSV 파일 URL (GCS Signed URL 등)
            table_name: 테이블 이름
            has_headers: 헤더 행 존재 여부
            delimiter: 구분자
            output_path: 출력 경로

        Returns:
            생성된 Excel 파일 경로
        """
        try:
            if output_path is None:
                output_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".xlsx"
                )
                output_path = output_file.name
                output_file.close()

            # Power Query M 코드 생성
            promote_headers = "Table.PromoteHeaders(Source, [PromoteAllScalars=true])" if has_headers else "Source"

            m_code = f'''let
    // CSV 파일을 웹에서 가져오기
    Source = Csv.Document(
        Web.Contents("{csv_url}"),
        [
            Delimiter="{delimiter}",
            Columns=null,
            Encoding=65001,
            QuoteStyle=QuoteStyle.None
        ]
    ),

    // 헤더 승격
    PromotedHeaders = {promote_headers}
in
    PromotedHeaders'''

            # Excel 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "연결 설정 방법"

            # 제목
            ws['A1'] = f"{table_name} - CSV 웹 데이터 연결"
            ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws.merge_cells('A1:B1')

            # CSV URL
            ws['A3'] = "데이터 소스:"
            ws['A3'].font = Font(bold=True)
            ws['B3'] = csv_url
            ws['B3'].alignment = Alignment(wrap_text=True)

            # 간단한 안내
            ws['A5'] = "🚀 빠른 연결 방법"
            ws['A5'].font = Font(size=12, bold=True)

            ws['A6'] = "1. 데이터 탭 → 데이터 가져오기 → 빈 쿼리"
            ws['A7'] = "2. 고급 편집기 열기"
            ws['A8'] = "3. 아래 'M 코드' 시트의 코드 복사/붙여넣기"
            ws['A9'] = "4. 완료 → 닫기 및 로드"

            ws['A11'] = "💡 이후 사용"
            ws['A11'].font = Font(size=11, bold=True)
            ws['A12'] = "• 데이터 새로고침: 데이터 탭 → 모두 새로고침"
            ws['A13'] = "• 연결은 파일에 저장되므로 다시 설정할 필요 없음"

            ws.column_dimensions['A'].width = 60
            ws.column_dimensions['B'].width = 50

            # M 코드 시트
            ws_code = wb.create_sheet(title="M 코드")
            ws_code['A1'] = "📝 아래 코드를 전체 선택하여 복사하세요"
            ws_code['A1'].font = Font(size=12, bold=True)

            ws_code['A3'] = m_code
            ws_code['A3'].font = Font(name="Consolas", size=10)
            ws_code['A3'].alignment = Alignment(wrap_text=True, vertical="top")
            ws_code['A3'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            ws_code.column_dimensions['A'].width = 100
            ws_code.row_dimensions[3].height = 150

            # 파일 저장
            wb.save(output_path)

            logger.info(f"Generated CSV connection Excel: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating CSV connection Excel: {str(e)}", exc_info=True)
            raise
