"""
Excel with Connection Service
Power Query 연결 정보만 포함된 Excel 파일 생성 (데이터는 포함하지 않음)
"""
import logging
import tempfile
import pandas as pd
from pathlib import Path
from typing import Optional, Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import zipfile
import shutil
import xml.etree.ElementTree as ET

logger = logging.getLogger(__name__)


class ExcelWithConnectionGenerator:
    """
    Power Query 연결이 설정된 Excel 파일 생성
    데이터는 시트에 포함하지 않고 연결 정보만 저장
    """

    def generate_excel_with_connection(
        self,
        odata_url: str,
        table_name: str,
        query_params: Optional[Dict[str, Any]] = None,
        output_path: Optional[str] = None
    ) -> str:
        """
        Power Query 연결이 포함된 Excel 파일 생성

        Args:
            odata_url: OData 서비스 URL
            table_name: 테이블/쿼리 이름
            query_params: OData 쿼리 파라미터 (filter, select, orderby 등)
            output_path: 출력 파일 경로

        Returns:
            생성된 Excel 파일 경로
        """
        try:
            if output_path is None:
                output_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".xlsx"
                )
                output_path = output_file.name
                output_file.close()

            # 기본 Excel 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            # 안내 메시지 추가
            ws['A1'] = f"OData 연결: {table_name}"
            ws['A1'].font = Font(size=14, bold=True)

            ws['A3'] = "이 통합 문서에는 OData 연결이 포함되어 있습니다."
            ws['A4'] = "데이터를 로드하려면:"
            ws['A5'] = "1. 데이터 탭 → 쿼리 및 연결 패널 열기"
            ws['A6'] = f"2. '{table_name}' 쿼리를 마우스 오른쪽 클릭"
            ws['A7'] = "3. '로드 대상...' 선택"
            ws['A8'] = "4. 원하는 위치 선택 후 '로드'"

            ws['A10'] = "또는 '모두 새로고침' 버튼을 클릭하여 데이터를 새로고침할 수 있습니다."

            # 컬럼 너비 조정
            ws.column_dimensions['A'].width = 60

            # 임시로 Excel 파일 저장
            temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_excel.name)
            temp_excel.close()

            # Excel 파일에 Power Query 연결 추가
            self._add_power_query_connection(
                temp_excel.name,
                output_path,
                odata_url,
                table_name,
                query_params
            )

            # 임시 파일 삭제
            Path(temp_excel.name).unlink(missing_ok=True)

            logger.info(f"Generated Excel with connection: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel with connection: {str(e)}", exc_info=True)
            raise

    def _add_power_query_connection(
        self,
        input_path: str,
        output_path: str,
        odata_url: str,
        table_name: str,
        query_params: Optional[Dict[str, Any]] = None
    ):
        """
        Excel 파일에 Power Query 연결 정보 추가

        ZIP 구조를 수정하여 xl/connections.xml과 xl/queryTables/ 추가
        """
        try:
            # 쿼리 파라미터가 있으면 URL에 추가
            full_url = odata_url
            if query_params:
                params = []
                if query_params.get('filter'):
                    params.append(f"$filter={query_params['filter']}")
                if query_params.get('select'):
                    params.append(f"$select={query_params['select']}")
                if query_params.get('orderby'):
                    params.append(f"$orderby={query_params['orderby']}")
                if params:
                    full_url += "?" + "&".join(params)

            # Power Query M 코드 생성
            m_code = f'''let
    Source = OData.Feed("{full_url}", null, [Implementation="2.0"])
in
    Source'''

            # 임시 디렉토리에 압축 해제
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(input_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir)

            # xl/connections.xml 생성
            connections_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <connection id="1" keepAlive="1" name="{table_name}" description="OData connection to {table_name}" type="100" refreshedVersion="7" minRefreshableVersion="5" saveData="0">
        <dbPr connection="Provider=Microsoft.Mashup.OleDb.1;Data Source=$Workbook$;Location={table_name}" command="{table_name}" commandType="3"/>
        <olapPr sendLocale="1"/>
        <extLst>
            <ext uri="{{DE250136-89BD-433C-8126-D09CA5730AF9}}" xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main">
                <x15:connection id="{table_name}"/>
            </ext>
        </extLst>
    </connection>
</connections>'''

            # connections.xml 파일 저장
            connections_path = Path(temp_dir) / "xl" / "connections.xml"
            connections_path.write_text(connections_xml, encoding='utf-8')

            # xl/_rels/workbook.xml.rels 수정 (connections.xml 참조 추가)
            rels_path = Path(temp_dir) / "xl" / "_rels" / "workbook.xml.rels"
            if rels_path.exists():
                tree = ET.parse(rels_path)
                root = tree.getroot()

                # 네임스페이스
                ns = {'': 'http://schemas.openxmlformats.org/package/2006/relationships'}

                # 새 관계 추가
                new_rel = ET.SubElement(root, 'Relationship')
                new_rel.set('Id', 'rIdConn1')
                new_rel.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/connections')
                new_rel.set('Target', 'connections.xml')

                tree.write(rels_path, encoding='utf-8', xml_declaration=True)

            # customXml 디렉토리 생성 및 Power Query 정보 추가
            customxml_dir = Path(temp_dir) / "customXml"
            customxml_dir.mkdir(exist_ok=True)

            # Power Query Mashup 데이터 생성 (간단한 버전)
            item1_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<DataMashup xmlns="http://schemas.microsoft.com/DataMashup">
    <Metadata>
        <ContentType>DataMashup</ContentType>
        <Version>1.0</Version>
    </Metadata>
    <Content>
        <Query Name="{table_name}">
            <Formula><![CDATA[{m_code}]]></Formula>
        </Query>
    </Content>
</DataMashup>'''

            item1_path = customxml_dir / "item1.xml"
            item1_path.write_text(item1_xml, encoding='utf-8')

            # customXml/_rels 디렉토리 생성
            customxml_rels_dir = customxml_dir / "_rels"
            customxml_rels_dir.mkdir(exist_ok=True)

            # workbook.xml 수정 (쿼리 테이블 참조 추가)
            workbook_path = Path(temp_dir) / "xl" / "workbook.xml"
            if workbook_path.exists():
                tree = ET.parse(workbook_path)
                root = tree.getroot()

                # 네임스페이스 등록
                ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                ET.register_namespace('', ns[''])

                # definedNames 요소 추가 (없으면)
                defined_names = root.find('.//{{{}}}definedNames'.format(ns['']))
                if defined_names is None:
                    defined_names = ET.SubElement(root, 'definedNames')

                # 쿼리 이름 정의 추가
                defined_name = ET.SubElement(defined_names, 'definedName')
                defined_name.set('name', f'_{table_name}')
                defined_name.set('hidden', '1')
                defined_name.text = f"'{table_name}'!$A$1"

                tree.write(workbook_path, encoding='utf-8', xml_declaration=True)

            # 수정된 파일들을 새 ZIP으로 압축
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
                for root_dir, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = Path(root_dir) / file
                        arcname = file_path.relative_to(temp_dir)
                        zip_out.write(file_path, arcname)

            # 임시 디렉토리 삭제
            shutil.rmtree(temp_dir)

        except Exception as e:
            logger.error(f"Error adding Power Query connection: {str(e)}", exc_info=True)
            # 임시 디렉토리가 있으면 삭제
            if 'temp_dir' in locals():
                shutil.rmtree(temp_dir, ignore_errors=True)
            raise


import os  # os 모듈 import 추가