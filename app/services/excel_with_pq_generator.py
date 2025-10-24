"""
Excel with Power Query Connection Generator
Power Query 연결이 이미 설정된 Excel 파일 생성 (간단한 DataMashup 사용)
"""
import logging
import tempfile
import zipfile
import shutil
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import base64
import uuid

logger = logging.getLogger(__name__)


class ExcelWithPowerQueryGenerator:
    """
    Power Query 연결이 설정된 Excel 파일 생성

    - 간단한 DataMashup 사용 (중첩 ZIP 없음)
    - workbook.xml에 dataModel 참조 추가
    - "쿼리 및 연결"에 쿼리가 표시됨
    - 사용자가 새로고침으로 데이터 로드
    """

    def generate_excel_with_power_query(
        self,
        json_api_url: str,
        table_name: str,
        query_name: Optional[str] = None,
        output_path: Optional[str] = None
    ) -> str:
        """
        Power Query 연결이 설정된 Excel 파일 생성

        Args:
            json_api_url: JSON API URL
            table_name: 테이블 이름
            query_name: 쿼리 이름 (기본: table_name)
            output_path: 출력 경로

        Returns:
            생성된 Excel 파일 경로
        """
        try:
            if output_path is None:
                output_file = tempfile.NamedTemporaryFile(
                    delete=False,
                    suffix=".xlsx"
                )
                output_path = output_file.name
                output_file.close()

            if query_name is None:
                query_name = table_name

            connection_name = f"쿼리 - {query_name}"

            # Power Query M 코드 생성
            m_code_text = f'''let
    Source = Json.Document(Web.Contents("{json_api_url}")),
    value = Source[value],
    ToTable = Table.FromRecords(value)
in
    ToTable'''

            # 1. 기본 Excel 파일 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            ws['A1'] = f"{table_name} - Power Query 연결"
            ws['A1'].font = Font(size=14, bold=True)

            ws['A3'] = "✅ 이 파일에는 Power Query 연결이 설정되어 있습니다."
            ws['A4'] = ""
            ws['A5'] = "데이터 로드 방법:"
            ws['A6'] = "  1. 데이터 탭 클릭"
            ws['A7'] = "  2. '쿼리 및 연결' 클릭"
            ws['A8'] = f"  3. '{query_name}' 쿼리를 마우스 오른쪽 클릭"
            ws['A9'] = "  4. '로드 대상...' 선택"
            ws['A10'] = "  5. 원하는 위치 선택 후 '로드'"
            ws['A11'] = ""
            ws['A12'] = "또는 데이터 탭 → '모두 새로고침' 클릭"
            ws['A13'] = ""
            ws['A14'] = f"📊 데이터 소스: {json_api_url}"

            ws.column_dimensions['A'].width = 60

            # 임시 파일로 저장
            temp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_xlsx.name)
            temp_xlsx.close()

            # 2. ZIP 압축 해제
            temp_dir = tempfile.mkdtemp()
            with zipfile.ZipFile(temp_xlsx.name, 'r') as z:
                z.extractall(temp_dir)

            # 3. xl/connections.xml 생성
            connection_uid = str(uuid.uuid4()).upper()
            connections_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" mc:Ignorable="xr16" xmlns:xr16="http://schemas.microsoft.com/office/spreadsheetml/2017/revision16">
    <connection id="1" xr16:uid="{{{connection_uid}}}" name="{connection_name}" description="Power Query connection to {query_name}" type="100" refreshedVersion="7" minRefreshableVersion="5" background="0">
        <extLst>
            <ext uri="{{DE250136-89BD-433C-8126-D09CA5730AF9}}" xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main">
                <x15:connection id="{query_name}"/>
            </ext>
        </extLst>
    </connection>
</connections>'''

            connections_path = Path(temp_dir) / "xl" / "connections.xml"
            connections_path.write_text(connections_xml, encoding='utf-8')

            logger.debug("Created connections.xml")

            # 4. customXml 디렉토리 및 간단한 DataMashup 생성
            customxml_dir = Path(temp_dir) / "customXml"
            customxml_dir.mkdir(exist_ok=True)

            # 간단한 Mashup XML (중첩 ZIP 없음)
            mashup_xml = f'''<Mashup xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns="http://schemas.microsoft.com/DataMashup">
<Client>EXCEL</Client>
<Version>2.116.622.0</Version>
<MinVersion>2.21.0.0</MinVersion>
<Culture>en-US</Culture>
<SafeCombine>false</SafeCombine>
<Items>
<Query Name="{query_name}">
<Formula><![CDATA[{m_code_text}]]></Formula>
<IsParameterQuery xsi:nil="true" />
<IsDirectQuery xsi:nil="true" />
</Query>
</Items>
</Mashup>'''

            mashup_base64 = base64.b64encode(mashup_xml.encode('utf-8')).decode('ascii')

            item1_xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<DataMashup xmlns="http://schemas.microsoft.com/DataMashup">{mashup_base64}</DataMashup>'''

            (customxml_dir / "item1.xml").write_text(item1_xml, encoding='utf-8')

            # itemProps1.xml
            itemprops_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<ds:datastoreItem ds:itemID="{5B725DA8-6340-4833-9E83-50DF7A96D20F}" xmlns:ds="http://schemas.openxmlformats.org/officeDocument/2006/customXml">
<ds:schemaRefs>
<ds:schemaRef ds:uri="http://schemas.microsoft.com/DataMashup"/>
</ds:schemaRefs>
</ds:datastoreItem>'''

            (customxml_dir / "itemProps1.xml").write_text(itemprops_xml, encoding='utf-8')

            # customXml/_rels/item1.xml.rels
            customxml_rels = customxml_dir / "_rels"
            customxml_rels.mkdir(exist_ok=True)

            rels_xml = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXmlProps" Target="itemProps1.xml"/>
</Relationships>'''

            (customxml_rels / "item1.xml.rels").write_text(rels_xml, encoding='utf-8')

            logger.debug("Created simple DataMashup (no nested ZIP)")

            # 5. xl/workbook.xml 수정 - dataModel 추가
            workbook_path = Path(temp_dir) / "xl" / "workbook.xml"
            tree = ET.parse(workbook_path)
            root = tree.getroot()

            # 네임스페이스 등록
            namespaces = {
                '': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
                'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
                'x15': 'http://schemas.microsoft.com/office/spreadsheetml/2010/11/main'
            }

            for prefix, uri in namespaces.items():
                if prefix:
                    ET.register_namespace(prefix, uri)
                else:
                    ET.register_namespace('', uri)

            # extLst 찾기 또는 생성
            extLst = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}extLst')
            if extLst is None:
                extLst = ET.SubElement(root, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}extLst')

            # dataModel ext 추가
            model_id = str(uuid.uuid4())
            ext_datamodel = ET.SubElement(extLst, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}ext')
            ext_datamodel.set('uri', '{FCE2AD5D-F65C-4FA6-A056-5C36A1767C68}')

            dataModel = ET.SubElement(ext_datamodel, '{http://schemas.microsoft.com/office/spreadsheetml/2010/11/main}dataModel')
            modelTables = ET.SubElement(dataModel, '{http://schemas.microsoft.com/office/spreadsheetml/2010/11/main}modelTables')

            modelTable = ET.SubElement(modelTables, '{http://schemas.microsoft.com/office/spreadsheetml/2010/11/main}modelTable')
            modelTable.set('id', f'{query_name}_{model_id}')
            modelTable.set('name', query_name)
            modelTable.set('connection', connection_name)

            tree.write(workbook_path, encoding='utf-8', xml_declaration=True)

            logger.debug("Added dataModel to workbook.xml")

            # 6. xl/_rels/workbook.xml.rels 수정
            rels_path = Path(temp_dir) / "xl" / "_rels" / "workbook.xml.rels"
            tree = ET.parse(rels_path)
            root = tree.getroot()

            ET.register_namespace('', 'http://schemas.openxmlformats.org/package/2006/relationships')

            # connections 관계 추가
            conn_rel = ET.SubElement(root, '{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
            conn_rel.set('Id', 'rIdConn1')
            conn_rel.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/connections')
            conn_rel.set('Target', 'connections.xml')

            tree.write(rels_path, encoding='utf-8', xml_declaration=True)

            # 7. _rels/.rels 수정 (customXml 참조)
            root_rels_path = Path(temp_dir) / "_rels" / ".rels"
            tree = ET.parse(root_rels_path)
            root = tree.getroot()

            custom_rel = ET.SubElement(root, '{http://schemas.openxmlformats.org/package/2006/relationships}Relationship')
            custom_rel.set('Id', 'rIdCustom1')
            custom_rel.set('Type', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/customXml')
            custom_rel.set('Target', 'customXml/item1.xml')

            tree.write(root_rels_path, encoding='utf-8', xml_declaration=True)

            # 8. [Content_Types].xml 수정
            content_types_path = Path(temp_dir) / "[Content_Types].xml"
            tree = ET.parse(content_types_path)
            root = tree.getroot()

            ET.register_namespace('', 'http://schemas.openxmlformats.org/package/2006/content-types')

            # connections.xml
            override1 = ET.SubElement(root, '{http://schemas.openxmlformats.org/package/2006/content-types}Override')
            override1.set('PartName', '/xl/connections.xml')
            override1.set('ContentType', 'application/vnd.openxmlformats-officedocument.spreadsheetml.connections+xml')

            # customXml/item1.xml
            override2 = ET.SubElement(root, '{http://schemas.openxmlformats.org/package/2006/content-types}Override')
            override2.set('PartName', '/customXml/item1.xml')
            override2.set('ContentType', 'application/xml')

            # customXml/itemProps1.xml
            override3 = ET.SubElement(root, '{http://schemas.openxmlformats.org/package/2006/content-types}Override')
            override3.set('PartName', '/customXml/itemProps1.xml')
            override3.set('ContentType', 'application/vnd.openxmlformats-officedocument.customXmlProperties+xml')

            tree.write(content_types_path, encoding='utf-8', xml_declaration=True)

            # 9. 다시 ZIP으로 압축
            with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for file_path in Path(temp_dir).rglob('*'):
                    if file_path.is_file():
                        arcname = file_path.relative_to(temp_dir)
                        zf.write(file_path, arcname)

            # 10. 정리
            Path(temp_xlsx.name).unlink()
            shutil.rmtree(temp_dir)

            logger.info(f"Generated Excel with Power Query connection: {output_path}")
            logger.info(f"Query: {query_name}, API: {json_api_url}")

            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel with Power Query: {str(e)}", exc_info=True)
            # 정리
            if 'temp_xlsx' in locals():
                Path(temp_xlsx.name).unlink(missing_ok=True)
            if 'temp_dir' in locals():
                shutil.rmtree(temp_dir, ignore_errors=True)
            raise
